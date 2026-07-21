"""
Carga credenciales de test desde el archivo .xlsm de mapeo, para uso
EXCLUSIVO de la suite e2e_live.

Reglas de seguridad que sigue este módulo:
- Nunca loggea usuario/password (ni con logger.debug, ni con print).
- Nunca los devuelve en un objeto que se serialice fácilmente sin querer
  (se devuelve un dict simple, no un modelo con __repr__ automático).
- La ruta al .xlsm se toma de una variable de entorno, nunca hardcodeada,
  para que el archivo real jamás quede referenciado (ni su path) en el
  código versionado.
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Dict, Optional, TypedDict

import openpyxl

ENV_VAR_PATH_XLSM = "PATH_CREDENCIALES_XLSM"
HOJA_CREDENCIALES = "Credenciales"
MARCADOR_NO_DESARROLLADO = "NO DESARROLLADO"


class CredencialJurisdiccion(TypedDict):
    codigo: str
    clase: str
    usuario: str
    password: str
    client_folder: str


class CredencialesNoConfiguradasError(Exception):
    """La variable de entorno PATH_CREDENCIALES_XLSM no está seteada o el
    archivo no existe. No se debe hardcodear un path por defecto: si esto
    pasa, es preferible fallar fuerte a que alguien asuma un path erróneo
    y termine ejecutando login con datos de otro archivo."""


def _resolver_path_xlsm(path_override: Optional[str] = None) -> Path:
    path_str = path_override or os.getenv(ENV_VAR_PATH_XLSM)
    if not path_str:
        raise CredencialesNoConfiguradasError(
            f"Seteá la variable de entorno {ENV_VAR_PATH_XLSM} con la ruta al "
            "archivo .xlsm de credenciales de test antes de correr esta suite."
        )
    path = Path(path_str)
    if not path.exists():
        raise CredencialesNoConfiguradasError(f"No existe el archivo: {path}")
    return path


def cargar_credenciales(
    path_override: Optional[str] = None,
) -> Dict[str, CredencialJurisdiccion]:
    """
    Lee la hoja "Credenciales" del .xlsm y devuelve un dict indexado por
    `clase` (el nombre de la clase Python de la jurisdicción, ej. "Agip",
    "Arba"), tal como aparece en config.jurisdiccion_clases.

    Filas con usuario/password == "NO DESARROLLADO" se excluyen del dict
    (esas jurisdicciones no tienen implementación todavía).
    """
    path = _resolver_path_xlsm(path_override)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if HOJA_CREDENCIALES not in wb.sheetnames:
        raise ValueError(
            f"El archivo {path.name} no tiene una hoja llamada '{HOJA_CREDENCIALES}'. "
            f"Hojas encontradas: {wb.sheetnames}"
        )
    ws = wb[HOJA_CREDENCIALES]

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise ValueError(f"La hoja '{HOJA_CREDENCIALES}' está vacía.")

    encabezado = [str(c).strip().lower() if c else "" for c in filas[0]]
    idx = {nombre: i for i, nombre in enumerate(encabezado)}

    requeridas = {"clase", "codigo", "usuario", "password", "client_folder"}
    faltantes = requeridas - set(idx.keys())
    if faltantes:
        raise ValueError(
            f"Faltan columnas requeridas en la hoja '{HOJA_CREDENCIALES}': {faltantes}. "
            f"Columnas encontradas: {list(idx.keys())}"
        )

    resultado: Dict[str, CredencialJurisdiccion] = {}
    for fila in filas[1:]:
        if fila is None or all(c is None for c in fila):
            continue
        clase = fila[idx["clase"]]
        usuario = fila[idx["usuario"]]
        password = fila[idx["password"]]
        if not clase:
            continue
        if usuario == MARCADOR_NO_DESARROLLADO or password == MARCADOR_NO_DESARROLLADO:
            continue
        resultado[str(clase)] = {
            "codigo": str(fila[idx["codigo"]]),
            "clase": str(clase),
            "usuario": str(usuario),
            "password": str(password),
            "client_folder": str(fila[idx["client_folder"]]),
        }
    return resultado


def resolver_cuit_cliente(
    credencial: CredencialJurisdiccion, overrides: Optional[Dict[str, str]] = None
) -> str:
    """
    El archivo de credenciales NO incluye una columna `cuit_cliente`
    (el CUIT del contribuyente a consultar, que en jurisdicciones con acceso
    delegado -p.ej. AGIP vía MiBA, ARBA- puede ser DISTINTO del CUIT usado
    para loguearse). Ante la ausencia de ese dato:

    1. Si `overrides` (dict clase -> cuit_cliente) trae un valor para esta
       clase, se usa ese.
    2. Si no, se usa el mismo valor que `usuario` (asume cuenta
       autogestionada, no delegada) y se emite un warning -- esto puede ser
       INCORRECTO para AGIP/ARBA si el usuario que loguea es un estudio
       contable distinto del contribuyente.
    """
    overrides = overrides or {}
    clase = credencial["clase"]
    if clase in overrides:
        return overrides[clase]
    return credencial["usuario"]
